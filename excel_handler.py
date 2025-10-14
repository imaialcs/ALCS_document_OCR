import sys
import json
import os
import re
import datetime
import traceback
from openpyxl.utils import column_index_from_string

# --- Logging Setup ---
_current_log_file_path = None

def set_log_file_path(log_dir=None):
    global _current_log_file_path
    if log_dir:
        _current_log_file_path = os.path.join(log_dir, "excel_handler_error.log")
    else:
        if getattr(sys, 'frozen', False):
            _script_dir = os.path.dirname(sys.executable)
        else:
            _script_dir = os.path.dirname(os.path.abspath(__file__))
        _current_log_file_path = os.path.join(_script_dir, "excel_handler_error.log")

def log_error(e, context=""):
    if not _current_log_file_path:
        print(f"LOGGING_ERROR: Log path not set. Original error: {context} - {e}", file=sys.stderr)
        return
    try:
        with open(_current_log_file_path, "a", encoding="utf-8") as log_file:
            log_file.write(f"[{datetime.datetime.now()}] Error {context}: {e}\n")
            log_file.write(traceback.format_exc())
            log_file.write("-" * 50 + "\n")
    except Exception as log_e:
        sys.__stderr__.write(f"CRITICAL ERROR: Failed to write to log file {log_e}\nOriginal Error {context}: {e}\n")

def global_exception_handler(exc_type, exc_value, exc_traceback):
    log_error(exc_value, f"Unhandled Exception: {exc_type.__name__}")
    sys.__excepthook__(exc_type, exc_value, exc_traceback)

sys.excepthook = global_exception_handler

# --- Library Imports ---
try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    log_error(e, "ImportError: openpyxl")
    OPENPYXL_AVAILABLE = False

def cell_to_row_col(cell_address):
    match = re.match(r"([A-Z]+)(\d+)", cell_address.upper())
    if not match:
        raise ValueError(f"Invalid cell address format: {cell_address}")
    col_name, row_str = match.groups()
    row_num = int(row_str)
    col_num = 0
    for char in col_name:
        col_num = col_num * 26 + (ord(char) - ord('A') + 1)
    return row_num, col_num

def find_sheet_safely(wb, target_sheet_name):
    if not target_sheet_name:
        return None
    if target_sheet_name in wb.sheetnames:
        return wb[target_sheet_name]
    return None

def read_roster_data(file_path, sheet_name, column_or_range, has_header=False):
    log_error(f"--- Starting read_roster_data with file: {file_path}, sheet: {sheet_name}, column: {column_or_range}, has_header: {has_header} ---", "INFO")
    if not OPENPYXL_AVAILABLE:
        return {"success": False, "error": "openpyxl library is not available."}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)
        sheet = find_sheet_safely(wb, sheet_name)
        if not sheet:
            raise ValueError(f"Sheet '{sheet_name or 'default'}' not found.")

        log_error(f"Successfully found sheet: {sheet.title}", "INFO")

        names = []
        # 氏名として除外するパターン
        EXCLUDE_PATTERNS = [
            re.compile(r'^\d+$'), # 純粋な数字
            re.compile(r'^\s*$'), # 空白のみ
            re.compile(r'従業員番号'), # 従業員番号
            re.compile(r'氏\s*名'), # 氏名ヘッダー
            re.compile(r'医療法人社団東鳩会'), # 医療法人社団東鳩会
            re.compile(r'^\d{4}-\d{2}-\d{2}.*$'), # 日付形式 (例: 1962-06-08 00:00:00)
        ]

        if ":" in column_or_range:
            rows = sheet[column_or_range]
            if has_header and len(rows) > 0:
                rows = rows[1:] # ヘッダーをスキップ
            for row in rows:
                for cell in row:
                    if cell.value is not None:
                        value_str = str(cell.value).strip()
                        # 氏名として適切かフィルタリング
                        is_excluded = False
                        for pattern in EXCLUDE_PATTERNS:
                            if pattern.search(value_str):
                                is_excluded = True
                                break
                        if value_str and not is_excluded:
                            names.append(value_str)
        else:
            col_idx = column_index_from_string(column_or_range.upper())
            start_row = 2 if has_header else 1

            consecutive_empty_cells = 0
            for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=start_row, max_row=sheet.max_row + 20): # Add buffer
                cell = row[0]
                if cell.value is not None and str(cell.value).strip() != "":
                    consecutive_empty_cells = 0 # Reset counter
                    value_str = str(cell.value).strip()
                    is_excluded = False
                    for pattern in EXCLUDE_PATTERNS:
                        if pattern.search(value_str):
                            is_excluded = True
                            break
                    if not is_excluded:
                        names.append(value_str)
                else:
                    consecutive_empty_cells += 1
                    if consecutive_empty_cells >= 20:
                        break # Stop if we hit 20 consecutive empty cells
        
        log_error(f"Finished reading. Found names: {names}", "INFO")
        return {"success": True, "names": names}
    except Exception as e:
        log_error(e, f"while reading roster data from {file_path}")
        return {"success": False, "error": str(e)}

def process_with_openpyxl(template_path, operations, output_path=None):
    if not OPENPYXL_AVAILABLE:
        return {"success": False, "error": "openpyxl library is not available."}
    is_xlsm = template_path.lower().endswith('.xlsm')
    wb = None
    try:
        wb = openpyxl.load_workbook(template_path, keep_vba=is_xlsm)
        processed_count = 0
        for op in operations:
            sheet_name = op.get("sheet_name")
            data = op.get("data")
            start_cell = op.get("start_cell", "E6")
            column_offsets = op.get("column_offsets", [])

            if not all([sheet_name, data, start_cell]):
                continue

            sheet = find_sheet_safely(wb, sheet_name)
            if not sheet:
                log_error(f"Sheet '{sheet_name}' not found in workbook.", "sheet not found")
                continue
            
            if sheet.protection.sheet:
                try:
                    sheet.protection.disable()
                except Exception as e:
                    log_error(e, f"while disabling sheet protection for {sheet_name}")
                    
            start_row, start_col = cell_to_row_col(start_cell)
            for i, row_data in enumerate(data):
                current_write_col = start_col
                for j, cell_value in enumerate(row_data):
                    sheet.cell(row=start_row + i, column=current_write_col, value=cell_value)
                    current_write_col += 1
                    if j < len(column_offsets):
                        current_write_col += column_offsets[j]
            processed_count += 1

        if processed_count > 0:
            save_path = output_path if output_path else template_path
            wb.save(save_path)
        return {"success": True, "message": f"{processed_count}件の操作をExcelファイルに正常に転記しました。"}
    except Exception as e:
        log_error(e, f"while processing with openpyxl: {template_path}")
        return {"success": False, "error": str(e)}

if __name__ == "__main__":
    set_log_file_path()
    result = {}
    try:
        if len(sys.argv) < 3:
            raise ValueError("Usage: excel_handler.py <input_json_path> <output_json_path>")

        json_file_path = sys.argv[1]
        output_file_path = sys.argv[2]

        with open(json_file_path, 'r', encoding='utf-8') as f:
            input_data = json.load(f)

        log_dir = input_data.get("log_dir")
        if log_dir:
            set_log_file_path(log_dir)

        action = input_data.get("action")

        if action == "read_roster":
            file_path = input_data.get("file_path")
            sheet_name = input_data.get("sheet_name", "リスト")
            column = input_data.get("column", "D")
            has_header = input_data.get("has_header", False) # has_headerを追加
            if not all([file_path, column]):
                raise ValueError("Missing required parameters for read_roster action.")
            result = read_roster_data(file_path, sheet_name, column, has_header)
        elif action == "write_template":
            template_path = input_data.get("template_path")
            operations = input_data.get("operations")
            output_path = input_data.get("output_path") # output_path を受け取る
            if not all([template_path, operations]):
                raise ValueError("Missing required parameters for write_template action.")
            result = process_with_openpyxl(template_path, operations, output_path)
        else:
            result = {"success": False, "error": f"Unknown action: {action}"}

    except Exception as e:
        log_error(e, "in main execution block")
        result = {"success": False, "error": str(e)}
    finally:
        try:
            with open(output_file_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False)
        except Exception as e:
            log_error(e, f"Failed to write result to output file {output_file_path}")
        
        if os.path.exists(json_file_path):
            try:
                os.remove(json_file_path)
            except OSError as e:
                log_error(e, f"Failed to remove temp file {json_file_path}")
